"""
Экспорт/импорт товаров через Excel (base64).
POST { action: "import", file_b64 } — импортирует товары из Excel (base64)
GET  — экспортирует все товары в Excel (base64)
"""
import json
import os
import base64
import psycopg2

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}

def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])

def ok(body, status=200):
    return {"statusCode": status, "headers": {**CORS, "Content-Type": "application/json"},
            "body": json.dumps(body, ensure_ascii=False, default=str)}

def err(msg, status=400):
    return {"statusCode": status, "headers": {**CORS, "Content-Type": "application/json"},
            "body": json.dumps({"error": msg})}


def get_categories(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, name, slug FROM categories")
    rows = cur.fetchall()
    cur.close()
    by_name = {r[1].lower(): r[0] for r in rows}
    by_slug = {r[2].lower(): r[0] for r in rows}
    return by_name, by_slug


def upsert_product(conn, p: dict, category_id):
    cur = conn.cursor()
    cur.execute("SELECT id FROM products WHERE name = %s", (p["name"],))
    row = cur.fetchone()
    image_urls = p.get("image_urls") or ([] if not p.get("image_url") else [p["image_url"]])
    image_url = p.get("image_url") or (image_urls[0] if image_urls else None)
    specs = p.get("specs") or {}
    if row:
        cur.execute("""
            UPDATE products SET price=%s, old_price=%s, in_stock=%s,
            image_url=COALESCE(%s, image_url),
            image_urls=CASE WHEN %s::jsonb != '[]'::jsonb THEN %s::jsonb ELSE image_urls END,
            specs=COALESCE(NULLIF(%s,'{}')::jsonb, specs),
            category_id=COALESCE(%s, category_id)
            WHERE id=%s
        """, (
            p["price"], p.get("old_price"), p.get("in_stock", True),
            image_url,
            json.dumps(image_urls), json.dumps(image_urls),
            json.dumps(specs),
            category_id, row[0]
        ))
        cur.close()
        return "updated", row[0]
    else:
        cur.execute("""
            INSERT INTO products (name, description, price, old_price, image_url, image_urls,
            specs, in_stock, is_featured, sort_order, category_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (
            p["name"], p.get("description", ""),
            p["price"], p.get("old_price"),
            image_url, json.dumps(image_urls), json.dumps(specs),
            p.get("in_stock", True), False, 0, category_id
        ))
        new_id = cur.fetchone()[0]
        cur.close()
        return "created", new_id


def export_excel(conn) -> str:
    import io
    import openpyxl
    cur = conn.cursor()
    cur.execute("""
        SELECT p.name, p.price, p.old_price, p.description, p.in_stock, p.is_featured,
               p.sort_order, p.image_url, p.specs::text, c.name as cat_name
        FROM products p
        LEFT JOIN categories c ON c.id = p.category_id
        ORDER BY p.sort_order, p.id
    """)
    rows = cur.fetchall()
    cur.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(["Название", "Цена", "Старая цена", "Описание", "В наличии",
               "Рекомендуем", "Сортировка", "Фото URL", "Характеристики (JSON)", "Категория"])
    for r in rows:
        ws.append([
            r[0], float(r[1]) if r[1] else 0, float(r[2]) if r[2] else None,
            r[3] or "", "Да" if r[4] else "Нет", "Да" if r[5] else "Нет",
            r[6] or 0, r[7] or "", r[8] or "{}", r[9] or ""
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()


def import_excel(file_b64: str, conn) -> dict:
    import io
    import openpyxl
    data = base64.b64decode(file_b64)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    cat_by_name, cat_by_slug = get_categories(conn)
    created, updated, skipped = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name:
            skipped += 1
            continue
        try:
            price = float(str(row[1] or 0).replace(" ", "").replace(",", "."))
        except Exception:
            price = 0.0
        try:
            old_price = float(str(row[2]).replace(" ", "").replace(",", ".")) if row[2] else None
        except Exception:
            old_price = None
        in_stock = str(row[4] or "да").lower() not in ("нет", "no", "false", "0")
        image_url = str(row[7] or "").strip() or None
        try:
            specs = json.loads(str(row[8] or "{}"))
        except Exception:
            specs = {}
        cat_raw = str(row[9] or "").strip().lower()
        category_id = cat_by_name.get(cat_raw) or cat_by_slug.get(cat_raw)
        product = {
            "name": name, "price": price, "old_price": old_price,
            "description": str(row[3] or ""),
            "image_url": image_url,
            "image_urls": [image_url] if image_url else [],
            "in_stock": in_stock, "specs": specs,
        }
        action, _ = upsert_product(conn, product, category_id)
        if action == "created":
            created += 1
        else:
            updated += 1
    conn.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    conn = get_conn()
    try:
        if event.get("httpMethod") == "GET":
            file_b64 = export_excel(conn)
            return ok({"file_b64": file_b64, "filename": "products.xlsx"})

        body = json.loads(event.get("body") or "{}")
        action = body.get("action")

        if action == "import":
            file_b64 = body.get("file_b64", "")
            if not file_b64:
                return err("file_b64 обязателен")
            result = import_excel(file_b64, conn)
            return ok(result)

        return err("Неизвестное действие")
    finally:
        conn.close()